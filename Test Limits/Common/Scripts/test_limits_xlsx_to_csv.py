# xlsx_converter.py
import sys
import os
import glob
import pandas as pd
import numpy as np
import re
import openpyxl
import argparse
from decimal import Decimal, ROUND_HALF_UP
from collections import Counter
import shutil
import csv
from enum import Enum, auto

pd.set_option('future.no_silent_downcasting', True)

class DecodingProfile(Enum):
    NO_PROFILE = auto()
    RAMPAGE_PROFILE = auto()
    CATAPULT_PROFILE = auto()

decoding_profile = DecodingProfile.NO_PROFILE

DEBUG_MESSAGE = False
SIMILARITY_THRESHOLD = 0.5
MAX_CLEAN_DECIMALS = 3
VOLT_DECIMALS = 1
AMP_DECIMALS = 3
# Backwards-compatible alias: some older code referenced `max_check_decimals`
# Ensure that remains defined to avoid NameError if encountered.
max_check_decimals = MAX_CLEAN_DECIMALS

def calculate_text_similarity(text_a, text_b):
    words_a = set(re.findall(r'\w+', text_a.lower()))
    words_b = set(re.findall(r'\w+', text_b.lower()))
    if not words_a or not words_b:
        return 0.0
    intersection = words_a.intersection(words_b)
    union = words_a.union(words_b)
    return len(intersection) / len(union)


def format_numeric(x, decimals=None, quantizer=None, max_clean_decimals=None):
    """Unified numeric formatter.

    - If `decimals` is provided (int), the output is formatted with exactly
      that many decimal places (fixed-point).
    - Else if `quantizer` is provided (Decimal), value is quantized with it
      and returned without unnecessary trailing zeros (integers printed as
      whole numbers).
    - Otherwise, this performs artifact-cleaning using `max_clean_decimals`
      and returns a human-friendly string.

    Handles numpy floats, ints, Decimal and numeric strings. Returns '' for NaN.
    """
    if max_clean_decimals is None:
        max_clean_decimals = MAX_CLEAN_DECIMALS
    if pd.isna(x):
        return ''
    try:
        # Pre-round floats with long binary tails
        if isinstance(x, (float, np.floating)):
            d_tmp = Decimal(str(x))
            if d_tmp.as_tuple().exponent < -max_clean_decimals:
                x = round(float(x), max_clean_decimals)
        d = Decimal(str(x))
    except Exception:
        return str(x)

    # If explicit decimals requested -> fixed-format output
    if isinstance(decimals, int):
        try:
            quant = Decimal(1).scaleb(-decimals)
            dq = d.quantize(quant, rounding=ROUND_HALF_UP)
            if dq == 0:
                dq = dq.copy_abs()
            fmt = f"{{0:.{decimals}f}}"
            return fmt.format(float(dq))
        except Exception:
            return str(d)

    # If quantizer provided -> quantize and trim trailing zeros
    if quantizer is not None:
        try:
            dq = d.quantize(quantizer, rounding=ROUND_HALF_UP)
            if dq == 0:
                dq = dq.copy_abs()
            if dq == dq.to_integral_value():
                return str(int(dq))
            return format(dq.normalize(), 'f')
        except Exception:
            return str(d)

    # Default artifact-cleaning behavior
    try:
        if d.as_tuple().exponent < -max_clean_decimals:
            quant = Decimal(1).scaleb(-max_clean_decimals)
            d = d.quantize(quant, rounding=ROUND_HALF_UP)
    except Exception:
        pass
    if DEBUG_MESSAGE and d.as_tuple().exponent < -max_clean_decimals:
        print(f"[DEBUG format_numeric] x={repr(x)} type={type(x)} d_after={d} exponent={d.as_tuple().exponent}")
    if d == d.to_integral_value():
        return str(int(d))
    try:
        return format(d.normalize(), 'f')
    except Exception:
        return str(d)


def clean_numeric_for_display(x, max_clean_decimals=MAX_CLEAN_DECIMALS):
    # Backwards-compatible wrapper around format_numeric
    return format_numeric(x, max_clean_decimals=max_clean_decimals)

# Shared helpers for per-column numeric formatting used by both writer and verifier
numeric_like_re = re.compile(r'^[+-]?\d+(?:\.\d+)?$')

def get_cleaner_for_col(col_name):
    col_lower = str(col_name).lower()
    if '(v)' in col_lower or 'volt' in col_lower:
        req_dec = VOLT_DECIMALS
    elif '(a)' in col_lower or 'amp' in col_lower or 'current' in col_lower:
        req_dec = AMP_DECIMALS
    else:
        req_dec = None

    def _clean_cell_for_output(v):
        if pd.isna(v) or (isinstance(v, str) and v.strip() == ''):
            return ''
        if req_dec is not None and (isinstance(v, (int, float, np.floating)) or (isinstance(v, str) and numeric_like_re.match(v.strip()))):
            return format_numeric(v, decimals=req_dec)
        if isinstance(v, (int, float, np.floating)):
            return format_numeric(v)
        if isinstance(v, str) and numeric_like_re.match(v.strip()):
            return format_numeric(v.strip())
        return str(v)

    return _clean_cell_for_output

def parse_sheet_name_for_description(sheet_name):
    desc = re.sub(r"^\d+\.\d+\s*_?\s*", "", sheet_name)
    desc = desc.replace("_", " ")
    return desc.strip()

def generate_csv_filename(section_text):
    file_name = str(section_text)
    sanitized_name = re.sub(r"^Section\s", "", file_name, flags=re.IGNORECASE).strip()
    sanitized_name = re.sub(r"-\sCCA\sONLY", "", sanitized_name, flags=re.IGNORECASE)
    sanitized_name = re.sub(r"-\sLow\sPower", "", sanitized_name, flags=re.IGNORECASE)
    sanitized_name = re.sub(r"(\d+.\d+),", r"\1", sanitized_name)
    sanitized_name = sanitized_name.replace("<", "").replace(">", "").replace(":", "")
    sanitized_name = sanitized_name.replace("\"", "").replace("/", "").replace("\\", "")
    sanitized_name = sanitized_name.replace("|", "").replace("?", "").replace("*", "")
    table_name = sanitized_name.strip().rstrip(",").strip()
    final_combined_filename = f"Section {table_name}"
    #print(f"\n+-----------------------------------+\n")
    #print(f"CSV FILE NAME: {final_combined_filename}\n")
    number_pattern = r"(\d+\.\d+)"
    number_list = re.findall(number_pattern, table_name)
    desc_part = re.sub(number_pattern, "", table_name).strip()
    desc_list = [d.strip() for d in desc_part.split('-') if d.strip()]
    if DEBUG_MESSAGE:
        print(f"[DEBUG] Parsed Numbers: {number_list} (Count: {len(number_list)})")
        print(f"[DEBUG] Parsed Descriptions: {desc_list} (Count: {len(desc_list)})")
    # If multiple numbers are present, do NOT split the section name into
    # separate per-test CSVs. Treat the combined section as a single table name.
    if len(number_list) > 1:
        if DEBUG_MESSAGE:
            print(f"[DEBUG] Multiple numbers detected in '{file_name}'. Keeping combined section name and not splitting.")
        return final_combined_filename, table_name, [], []
    return final_combined_filename, table_name, number_list, desc_list

def duplicate_csv_for_each_test(output_folder, source_csv_name, number_list, desc_list):
    # Duplication of the combined CSV per test is disabled.
    # Always return an empty list to indicate no new files were created.
    if DEBUG_MESSAGE and number_list:
        print(f"[DEBUG] duplicate_csv_for_each_test called but duplication is disabled. number_list={number_list}, desc_list={desc_list}")
    return []

def find_xlsx_files():
    source_folder = "Test Limit Source Files"
    if not os.path.isdir(source_folder):
        print(f"Error: The source directory '{source_folder}' was not found.")
        input("--- Press Enter to continue ---")
        return []
    search_path = os.path.join(source_folder, "*.xlsx")
    all_files = glob.glob(search_path)
    return [f for f in all_files if not os.path.basename(f).startswith('~')]
 
def select_xlsx_file(file_list):
    if not file_list:
        print("No .xlsx files found.")
        input("--- Press Enter to continue ---")
        return None

    print("Would you like to convert these XLSX files to CSV?\n")
    print("Selection options:")
    print("  - Single number (e.g. 2)")
    print("  - Multiple numbers (e.g. 1,3,5)")
    print("  - Ranges (e.g. 1-4)")
    print("  - Mixed (e.g. 1-3,6)")
    
    print("\nAvailable L3 Test Limit files:")
    print("  0. None")

    for i, filename in enumerate(file_list, 1):
        basename = os.path.splitext(os.path.basename(filename))[0]
        exists = os.path.isdir(basename)
        print(f"  {i}. {os.path.basename(filename)}{' (already converted)' if exists else ''}")

    while True:
        choice = input("\nSelect file(s): ").strip()

        # None
        if choice == "0":
            return []

        try:
            selected_indices = set()

            for part in choice.split(","):
                part = part.strip()
                if "-" in part:
                    start, end = map(int, part.split("-"))
                    selected_indices.update(range(start, end + 1))
                else:
                    selected_indices.add(int(part))

            # Validate indices
            if not selected_indices:
                raise ValueError

            if any(i < 1 or i > len(file_list) for i in selected_indices):
                raise ValueError

            # Return list of selected files
            return [file_list[i - 1] for i in sorted(selected_indices)]

        except ValueError:
            print("Invalid selection. Please try again.")

def select_sheets(sheet_names):
    if not sheet_names:
        print("No sheets available in the selected file.")
        return []
    print("\n--- Available Sheets ---")
    print("  0. All Sheets")
    for i, name in enumerate(sheet_names, 1): print(f"  {i}. {name}")
    while True:
        try:
            choice = int(input("\nPlease select a sheet by number: "))
            if choice == 0: return sheet_names
            if 1 <= choice <= len(sheet_names): return [sheet_names[choice - 1]]
            else: print("Invalid selection.")
        except ValueError: print("Invalid input.")

def trim_empty_rows_and_cols(df):
    df_copy = df.copy()
    df_copy = df_copy.replace(r'[\n\r\t]+', ' ', regex=True).replace(r'\s+', ' ', regex=True)
    df_copy = df_copy.replace('', np.nan).dropna(how='all').dropna(how='all', axis=1)
    df_copy.reset_index(drop=True, inplace=True)
    df_copy.columns = range(df_copy.shape[1])
    return df_copy

def remerge_cells(filepath, sheet_name):
    if DEBUG_MESSAGE:
        print(f"  - Pre-processing sheet '{sheet_name}' to handle merged cells...")
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    sheet = workbook[sheet_name]
    merged_ranges = list(sheet.merged_cells.ranges)
    rows = list(sheet.iter_rows(values_only=True))
    df_raw = pd.DataFrame(rows)
    # For any merged cell range, keep only the top-left value and clear
    # all other cells in the merge region. This prevents duplicated
    # values when converting merged cells across columns (horizontal merges)
    # or down rows (vertical merges).
    for merged_range in merged_ranges:
        start_row, end_row = merged_range.min_row - 1, merged_range.max_row - 1
        start_col, end_col = merged_range.min_col - 1, merged_range.max_col - 1
        try:
            top_value = df_raw.iat[start_row, start_col]
        except Exception:
            top_value = None
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                if r == start_row and c == start_col:
                    df_raw.iat[r, c] = top_value
                else:
                    df_raw.iat[r, c] = None
    return df_raw

def _find_and_extract_one_table(df_slice, sheet_name):
    # Initialize a local profile for this specific table find operation.
    profile = DecodingProfile.NO_PROFILE

    first_valid_row_pos = 0
    if not df_slice.dropna(how='all').empty:
        first_valid_row_label = df_slice.dropna(how='all').index[0]
        first_valid_row_pos = df_slice.index.get_loc(first_valid_row_label)
    stripped_df = trim_empty_rows_and_cols(df_slice)
    if stripped_df.empty: return None, len(df_slice), None, profile
    # if DEBUG_MESSAGE and not stripped_df.empty:
    #     print(f"\n--- Analyzing Raw Data Block from sheet '{sheet_name}' ---")
    #     print(stripped_df.to_string())
    #     input("Press Enter to continue analysis...")

    section_cell_text = None
    anchor_row_index = -1
    header_row_index_in_stripped_df = -1

    section_mask = stripped_df[0].astype(str).str.match(r"Section\s", na=False)
    section_coords = np.where(section_mask)
    if section_coords[0].size > 0:
        anchor_row_index = section_coords[0][0]
        section_cell_text = stripped_df.iloc[anchor_row_index, 0]
        profile = DecodingProfile.RAMPAGE_PROFILE
    else:
        desc_text = parse_sheet_name_for_description(sheet_name)
        test_header_present = stripped_df.astype(str).apply(lambda x: x.str.contains('test|test number', na=False, case=False)).any().any()
        if test_header_present:
            for index, row in stripped_df.iterrows():
                row_as_string = ' '.join(row.dropna().astype(str))
                similarity = calculate_text_similarity(desc_text, row_as_string)
                contains_section = 'section' in row_as_string.lower()
                contains_test_number = re.search(r'\d+\.\d+', row_as_string)
                if similarity >= SIMILARITY_THRESHOLD and not contains_section and not contains_test_number:
                    section_cell_text = sheet_name
                    anchor_row_index = index
                    profile = DecodingProfile.CATAPULT_PROFILE
                    break
    if not section_cell_text:
        desc_text = parse_sheet_name_for_description(sheet_name)
        possible_header_rows = stripped_df[stripped_df.astype(str).apply(lambda r: r.str.contains('test|test number', na=False, case=False)).any(axis=1)]
        if not possible_header_rows.empty:
            for h_index in possible_header_rows.index:
                if h_index > 0 and stripped_df.shape[1] > 2:
                    disambiguation_row_index = h_index - 1
                    cell1 = str(stripped_df.iloc[disambiguation_row_index, 1] or '')
                    cell2 = str(stripped_df.iloc[disambiguation_row_index, 2] or '')
                    combined_text = f"{cell1} {cell2}".strip()
                    if combined_text and calculate_text_similarity(desc_text, combined_text) >= SIMILARITY_THRESHOLD:
                        section_cell_text = sheet_name
                        anchor_row_index = disambiguation_row_index
                        header_row_index_in_stripped_df = h_index
                        profile = DecodingProfile.CATAPULT_PROFILE
                        break
    if not section_cell_text: return None, len(df_slice), None, profile
    parsing_results = generate_csv_filename(section_cell_text)
    if header_row_index_in_stripped_df == -1:
        search_start_row = anchor_row_index + 1
        search_area_df = stripped_df.iloc[search_start_row:]
        if search_area_df.empty: return None, len(df_slice), None, profile
        search_df = search_area_df.astype(str).apply(lambda x: x.str.strip().str.lower())
        boolean_mask = search_df.apply(lambda x: x.str.contains('test|test number', na=False))
        coords = np.where(boolean_mask)
        if coords[0].size == 0: return None, len(df_slice), None, profile
        header_row_index_in_stripped_df = search_start_row + coords[0][0]
    start_col_index = stripped_df.iloc[header_row_index_in_stripped_df].dropna().index[0]
    original_headers = stripped_df.iloc[header_row_index_in_stripped_df].tolist()
    clean_headers_list = [str(h).strip() if pd.notna(h) else f'Unnamed_{i}' for i, h in enumerate(original_headers)]
    header_counts = Counter(h for h in clean_headers_list if not h.startswith('Unnamed'))
    duplicates = {h for h, c in header_counts.items() if c > 1}
    unique_headers = []
    if duplicates and header_row_index_in_stripped_df > 0:
        disambiguation_series = stripped_df.iloc[header_row_index_in_stripped_df - 1].copy()
        disambiguation_series.ffill(inplace=True)
        disambiguation_row = disambiguation_series.tolist()
        for i, header in enumerate(clean_headers_list):
            if header in duplicates:
                unique_headers.append(f"{str(disambiguation_row[i]).strip()} {header}")
            else:
                unique_headers.append(header)
    else:
        unique_headers = clean_headers_list
    counts = {}
    final_unique_headers = []
    for header in unique_headers:
        if header in counts:
            counts[header] += 1
            final_unique_headers.append(f"{header}_{counts[header]}")
        else:
            counts[header] = 1
            final_unique_headers.append(header)
    data_df = stripped_df.iloc[header_row_index_in_stripped_df + 1:, start_col_index:].copy()
    data_df.columns = final_unique_headers[start_col_index:]
    data_df.reset_index(drop=True, inplace=True)
    headers = data_df.columns.tolist()
    validated_data_rows = []
    rows_in_table_data = 0
    parsing_state = "READING_TESTS"
    for index, row in data_df.iterrows():
        first_cell_value = row[headers[0]]
        if pd.isna(first_cell_value) or str(first_cell_value).strip() == '': break
        first_cell_str = str(first_cell_value).strip()
        if first_cell_str.lower().startswith("section"): break
        is_test_row = first_cell_str.lower().startswith("test_")
        if parsing_state == "READING_TESTS":
            if is_test_row:
                validated_data_rows.append(row)
                rows_in_table_data += 1
            else:
                if validated_data_rows: parsing_state = "READING_NOTES"
                else: raise ValueError(f"Expected 'Test_', found '{first_cell_str}'.")
        elif parsing_state == "READING_NOTES":
            if is_test_row: raise ValueError(f"Found 'Test_' after a note: '{first_cell_str}'")
    if not validated_data_rows: return None, len(df_slice), None, profile
    final_df = pd.DataFrame(validated_data_rows, columns=headers)
    cols_to_drop = [col for col in final_df.columns if str(col).startswith('Unnamed_')]
    final_df = final_df.drop(columns=cols_to_drop)
    rows_consumed = first_valid_row_pos + header_row_index_in_stripped_df + 1 + rows_in_table_data
    return final_df, rows_consumed, parsing_results, profile

def extract_all_tables_from_sheet(raw_df, sheet_name):
    all_tables_data = []
    offset = 0
    while offset < len(raw_df):
        df_slice = raw_df.iloc[offset:]
        table_df, rows_consumed, parsing_results, profile = _find_and_extract_one_table(df_slice, sheet_name)
        if table_df is not None:
            all_tables_data.append(((table_df, parsing_results), profile))
        if not rows_consumed: break
        offset += rows_consumed
    
    num_catapult_profiles = sum(1 for _, profile in all_tables_data if profile == DecodingProfile.CATAPULT_PROFILE)
    if num_catapult_profiles > 0 and len(all_tables_data) > 1:
        print(f"[ERROR] On sheet '{sheet_name}', a table was found using CATAPULT_PROFILE, but more than one table was found in total.")
        print("        Sheets with Catapult profiles are only allowed to have one table.")
        input("--- Press Enter to continue ---")
        return []
    
    return all_tables_data

def process_sheets_to_csv(filepath, output_folder, sheets_to_process):
    processed_files_info = []
    for sheet_name in sheets_to_process:
        #print("\n+-----------------------------------+")
        print(f"\n--- Processing Sheet: '{sheet_name}' ---")
        raw_df = remerge_cells(filepath, sheet_name)
        extracted_tables = extract_all_tables_from_sheet(raw_df, sheet_name)
        if not extracted_tables:
            print("  - No tables found on this sheet.")
            input("--- Press Enter to continue ---")
            continue
        filename_counts = Counter()
        for i, ((final_df, (filename_base, table_name, number_list, desc_list)), profile) in enumerate(extracted_tables):
            for col in final_df.columns:
                if pd.api.types.is_bool_dtype(final_df[col]):
                    final_df[col] = final_df[col].astype(object)
                    final_df.loc[:, col] = final_df[col].apply(lambda x: str(x).upper() if pd.notna(x) else '')
                elif pd.api.types.is_numeric_dtype(final_df[col]):
                    final_df[col] = final_df[col].astype(object)
                    is_int = all(pd.isna(x) or (isinstance(x, (int, float)) and x == int(x)) for x in final_df[col])
                    if is_int:
                        final_df.loc[:, col] = final_df[col].apply(lambda x: f'{int(x):.0f}' if pd.notna(x) else '')
                    else:
                        # --- MODIFIED: Conditional rounding based on profile ---
                        if profile == DecodingProfile.RAMPAGE_PROFILE:
                            # No rounding for Rampage profile, but clean binary float artifacts
                            final_df.loc[:, col] = final_df[col].apply(lambda v: format_numeric(v))
                        else: # Catapult profile or default
                            # Choose quantizer based on column header: two decimals for specific current limit header
                            col_name = str(col).strip()
                            if "ext dc current limit" in col_name.lower():
                                quantizer = Decimal('0.00')
                            else:
                                quantizer = Decimal('0.0')
                            final_df.loc[:, col] = final_df[col].apply(lambda x: format_numeric(x, quantizer=quantizer))

            filename_counts[filename_base] += 1
            count = filename_counts[filename_base]
            suffix = f"_{count}" if count > 1 else ""
            csv_filename = f"{filename_base}{suffix}.csv"
            output_path = os.path.join(output_folder, csv_filename)
            source_xlsx_filename = os.path.basename(filepath)
            final_table_name = table_name if table_name.lower().startswith('section') else f"Section {table_name}"
            final_df_clean = final_df.copy()
            for col in final_df_clean.columns:
                cleaner = get_cleaner_for_col(col)
                final_df_clean.loc[:, col] = final_df_clean[col].apply(cleaner)

            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Source File", source_xlsx_filename])
                writer.writerow(["Table Name", final_table_name])
                writer.writerow([])
            final_df_clean.to_csv(output_path, mode='a', header=True, index=False)
            print(f"  - Successfully saved '{csv_filename}'")
            newly_created_files = duplicate_csv_for_each_test(output_folder, csv_filename, number_list, desc_list)
            if newly_created_files:
                for new_file in newly_created_files:
                    processed_files_info.append({"sheet_name": sheet_name, "csv_filename": new_file, "table_index": i})
            else:
                processed_files_info.append({"sheet_name": sheet_name, "csv_filename": csv_filename, "table_index": i})
    return processed_files_info

def verify_single_sheet(filepath, output_folder, sheet_name, processed_files_info_for_sheet):
    print(f"\n--- Verifying Sheet: '{sheet_name}' ---")
    any_failures_found = False
    if not processed_files_info_for_sheet:
        print("  - No files to verify for this sheet.")
        return False
    try:
        raw_df = remerge_cells(filepath, sheet_name)
        excel_tables_info = extract_all_tables_from_sheet(raw_df, sheet_name)
        for info in processed_files_info_for_sheet:
            is_file_verified = True
            if info['table_index'] >= len(excel_tables_info):
                print(f"  - ❌ FAILED for '{info['csv_filename']}': Cannot find matching source table in Excel.")
                any_failures_found = True
                input("--- Press Enter to continue ---")
                continue
            
            (excel_df, _), profile = excel_tables_info[info['table_index']]

            for col in excel_df.columns:
                if pd.api.types.is_bool_dtype(excel_df[col]):
                    excel_df[col] = excel_df[col].astype(object)
                    excel_df.loc[:, col] = excel_df[col].apply(lambda x: str(x).upper() if pd.notna(x) else '')
                elif pd.api.types.is_numeric_dtype(excel_df[col]):
                    excel_df[col] = excel_df[col].astype(object)
                    is_int = all(pd.isna(x) or (isinstance(x, (int, float)) and x == int(x)) for x in excel_df[col])
                    if is_int:
                        excel_df.loc[:, col] = excel_df[col].apply(lambda x: f'{int(x):.0f}' if pd.notna(x) else '')
                    else:
                        # --- MODIFIED: Conditional rounding based on profile ---
                        if profile == DecodingProfile.RAMPAGE_PROFILE:
                            # No rounding for Rampage profile, but clean binary float artifacts
                            excel_df.loc[:, col] = excel_df[col].apply(lambda v: format_numeric(v))
                        else: # Catapult or default
                            # Choose quantizer based on column header: two decimals for specific current limit header
                            col_name = str(col).strip()
                            if "ext dc current limit" in col_name.lower():
                                quantizer = Decimal('0.00')
                            else:
                                quantizer = Decimal('0.0')
                            excel_df.loc[:, col] = excel_df[col].apply(lambda x: format_numeric(x, quantizer=quantizer))

            # Apply shared cleaning to excel_df so verifier uses same formatting
            for col in excel_df.columns:
                cleaner = get_cleaner_for_col(col)
                excel_df.loc[:, col] = excel_df[col].apply(cleaner)
            excel_df = excel_df.fillna('').astype(str)
            csv_path = os.path.join(output_folder, info['csv_filename'])
            if not os.path.exists(csv_path):
                print(f"  - ❌ FAILED for '{info['csv_filename']}': CSV file not found.")
                any_failures_found = True
                input("--- Press Enter to continue ---")
                continue
            csv_df = pd.read_csv(csv_path, dtype=str, keep_default_na=False, skiprows=3)
            # Normalize CSV cell values using the shared per-column cleaner
            for col in csv_df.columns:
                cleaner = get_cleaner_for_col(col)
                csv_df[col] = csv_df[col].map(cleaner)
            if not excel_df.columns.equals(csv_df.columns):
                print(f"  - ❌ FAILED for '{info['csv_filename']}': Headers do not match.")
                any_failures_found = True
                is_file_verified = False
                input("--- Press Enter to continue ---")
            if is_file_verified and not excel_df.reset_index(drop=True).equals(csv_df.reset_index(drop=True)):
                print(f"  - ❌ FAILED for '{info['csv_filename']}': Data content does not match.")
                any_failures_found = True
                is_file_verified = False
                if DEBUG_MESSAGE:
                    print("\n--- EXPECTED DATA (from Excel) ---")
                    print(excel_df.to_string())
                    print("\n--- ACTUAL DATA (from CSV) ---")
                    print(csv_df.to_string())
                    print("\n" + "="*35)
                input("--- Press Enter to continue ---")
            if is_file_verified:
                print(f"  - ✅ SUCCESSFUL for '{info['csv_filename']}'.")
        return any_failures_found
    except Exception as e:
        print(f"\n[ERROR] An unexpected error occurred during verification of sheet '{sheet_name}': {e}")
        input("--- Press Enter to continue ---")
        return True

def main():
    global DEBUG_MESSAGE, MAX_CLEAN_DECIMALS, max_check_decimals
    parser = argparse.ArgumentParser(description="Convert XLSX sheets to CSV.")
    parser.add_argument('--input_path')
    parser.add_argument('--output_path')
    parser.add_argument('-a', '--all_sheets', action='store_true')
    parser.add_argument('-d', '--debug', action='store_true')
    parser.add_argument('--max-clean-decimals', type=int, default=MAX_CLEAN_DECIMALS,
                        help='Number of fractional decimals to keep when cleaning float artifacts (default: 3)')
    args = parser.parse_args()
    if args.debug:
        DEBUG_MESSAGE = True
    # Apply CLI override for max clean decimals
    if args.max_clean_decimals is not None:
        MAX_CLEAN_DECIMALS = int(args.max_clean_decimals)
        max_check_decimals = MAX_CLEAN_DECIMALS
    if args.input_path:
        files_to_process = [args.input_path]
    else:
        xlsx_files = find_xlsx_files()
        if not xlsx_files: return True
        selected_option = select_xlsx_file(xlsx_files)
        if not selected_option: return
        files_to_process = selected_option if isinstance(selected_option, list) else [selected_option]
    overall_verification_failed = False
    for i, filepath in enumerate(files_to_process):
        print(f"\n--- Processing file {i+1}/{len(files_to_process)}: {os.path.basename(filepath)} ---")
        try:
            xlsx = pd.ExcelFile(filepath)
            sheets_to_process = xlsx.sheet_names if args.all_sheets else select_sheets(xlsx.sheet_names)
            if not sheets_to_process:
                print("No sheets selected. Skipping file.")
                continue
            output_folder = os.path.splitext(os.path.basename(filepath))[0]
            if args.output_path: output_folder = os.path.join(args.output_path, output_folder)
            if os.path.exists(output_folder): shutil.rmtree(output_folder)
            os.makedirs(output_folder)
            print(f"\nCreated output folder: {output_folder}")
            all_processed_files = []
            for sheet_name in sheets_to_process:
                processed_files_info_for_sheet = process_sheets_to_csv(filepath, output_folder, [sheet_name])
                all_processed_files.extend(processed_files_info_for_sheet)
            if all_processed_files:
                print("\n--- Verification Starting for all processed sheets ---")
                from itertools import groupby
                key_func = lambda x: x['sheet_name']
                grouped_info = groupby(sorted(all_processed_files, key=key_func), key_func)
                for sheet_name, group in grouped_info:
                    if verify_single_sheet(filepath, output_folder, sheet_name, list(group)):
                        overall_verification_failed = True
        except Exception as e:
            print(f"\n[FATAL ERROR] processing {os.path.basename(filepath)}: {e}")
            input("--- Press Enter to continue ---")
            overall_verification_failed = True
        if len(files_to_process) > 1 and i < len(files_to_process) - 1:
            input("\nPress Enter to continue to the next file...")
    print("\n--- All Processing Complete ---")
    if overall_verification_failed:
        print("⚠️ Some files failed verification during the run.")
        input("\n--- Press Enter to continue ---")
        sys.exit()
    else:
        print("🎉 All files processed and verified successfully!")
        print("\n--- Reminder ---")
        print("Next step: Update TestLimitsLookup.csv with any new test limit directories and their associated hardware revisions.")
        input("\n--- Press Enter to continue ---")

if __name__ == "__main__":
    main()
